# SPDX-License-Identifier: Apache-2.0
"""
Gradio Chatbot Interface for vllm-mlx.

A multimodal chat interface that connects to the vllm-mlx server
and supports text, images, video, and documents (PDF, DOCX, TXT).

Usage:
    # First start the server with a multimodal model:
    vllm-mlx serve --served-model-name default mlx-community/Qwen3-VL-4B-Instruct-3bit --port 8000

    # Then run the app:
    vllm-mlx-chat

    # Or with a different served-model name served on localhost:8000:
    vllm-mlx-chat --served-model-name <served-model-name>  --server-url http://localhost:8000 --port 7860

Note:
    Query the /v1/models endpoint on localhost with `curl` and `jq` to see available models and their names:
    ```bash
    curl http://localhost:8000/v1/models | jq ".data[0].id"
    ```
"""

import argparse
import base64
import mimetypes
import shutil
import subprocess
import tempfile
from pathlib import Path

import gradio as gr
import requests

# Image MIME types by extension.
IMAGE_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
    ".heic": "image/heic",
    ".heif": "image/heif",
    ".tif": "image/tiff",
    ".tiff": "image/tiff",
}

# Video MIME types by extension. ".mov" is what an iPhone produces and
# "video/quicktime" is the MIME browsers report for it — an allow-list missing
# either entry drops the attachment with no error anywhere in the stack.
VIDEO_TYPES = {
    ".mp4": "video/mp4",
    ".m4v": "video/x-m4v",
    ".webm": "video/webm",
    ".mov": "video/quicktime",
    ".qt": "video/quicktime",
    ".avi": "video/x-msvideo",
    ".mkv": "video/x-matroska",
    ".mpeg": "video/mpeg",
    ".mpg": "video/mpeg",
    ".ogv": "video/ogg",
    ".3gp": "video/3gpp",
}

# Documents are not sent to the model as pixels. A PDF has a text layer, and
# reading it is both cheaper and far more accurate than asking a vision encoder
# to read 9pt type: a page of text costs ~1.2k tokens and is exact, where the
# same page rendered at 150 DPI costs ~2k tokens and invites misread digits.
# So a PDF becomes a *text* content part, which needs no server support at all.
DOCUMENT_TYPES = {
    ".pdf": "application/pdf",
    ".txt": "text/plain",
    ".md": "text/markdown",
    ".csv": "text/csv",
    ".docx": (
        "application/vnd.openxmlformats-officedocument"
        ".wordprocessingml.document"
    ),
    ".xlsx": (
        "application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet"
    ),
    ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
}

# Rejoining separator per chunk unit, so reassembled text reads the way the
# source did rather than running rows or pages together.
CHUNK_SEPARATORS = {"page": "\f", "paragraph": "\n\n", "row": "\n"}

# Rough chars-per-token for English prose. Only used to warn and to trim, so an
# approximation is fine — the server enforces the real limit.
CHARS_PER_TOKEN = 4

# Leave room in the request for the question and the model's answer. The server
# default is 32768 max request tokens.
MAX_DOCUMENT_TOKENS = 20000

# A PDF whose pages yield almost no characters is a scan: the text layer is
# absent and pdftotext has nothing to give. Say so rather than send empty text.
MIN_CHARS_PER_PAGE = 40

# Figures are sent as images alongside the text, because a chart's meaning is in
# the picture: a caption reading "Illegal = bars" describes a hatch pattern that
# cannot be recovered from the text layer, and a model given only the caption
# will confidently invent what the figure looks like.
MAX_FIGURE_IMAGES = 4

# Images cost pixels / (patch^2 * merge^2) = pixels / 1024 tokens on Qwen3-VL
# class models. Keep the figures well inside the request budget alongside text.
MAX_FIGURE_TOKENS = 9000
FIGURE_PIXELS_PER_TOKEN = 1024

# Below this, an embedded image is a logo, rule, or icon rather than a figure.
MIN_FIGURE_EDGE = 300

# Extensions accepted by the upload widget, listed explicitly alongside the
# "image"/"video" categories so a container the browser fails to type-sniff is
# still offered to the server rather than rejected in the file picker.
UPLOAD_FILE_TYPES = (
    ["image", "video"]
    + sorted(IMAGE_TYPES)
    + sorted(VIDEO_TYPES)
    + sorted(DOCUMENT_TYPES)
)


def _read_document(file_path: str) -> tuple[list[str], str]:
    """Read a document into chunks, with the name of what a chunk is.

    Chunks exist so truncation lands on a real boundary and can be reported in
    a unit the reader recognises: pages for a PDF, paragraphs for plain text.

    Raises:
        ValueError: If the format is unsupported or the file cannot be read.
    """
    path = Path(file_path)
    name, suffix = path.name, path.suffix.lower()

    if suffix == ".pdf":
        if shutil.which("pdftotext") is None:
            raise ValueError(
                f"Cannot read {name}: pdftotext is not installed. "
                "Install it with: brew install poppler"
            )
        text = _run_extractor(["pdftotext", "-layout", file_path, "-"], name)
        # pdftotext separates pages with a form feed.
        return [p for p in text.split("\f") if p.strip()], "page"

    if suffix == ".docx":
        if shutil.which("textutil") is None:
            raise ValueError(f"Cannot read {name}: textutil is unavailable.")
        text = _run_extractor(
            ["textutil", "-convert", "txt", "-stdout", file_path], name
        )
        return [p for p in text.split("\n\n") if p.strip()], "paragraph"

    if suffix in (".xlsx", ".xlsm"):
        return _read_spreadsheet(file_path, name), "row"

    if suffix == ".csv":
        return [ln for ln in _read_text_file(path).splitlines() if ln.strip()], "row"

    if suffix in (".txt", ".md"):
        text = _read_text_file(path)
        return [p for p in text.split("\n\n") if p.strip()], "paragraph"

    raise ValueError(f"Unsupported document type {suffix!r}")


def _read_text_file(path: Path) -> str:
    """Read a text file, tolerating non-UTF-8 bytes.

    latin-1 decodes any byte sequence, so a mangled character beats refusing
    the document outright.
    """
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="latin-1")
    except OSError as e:
        raise ValueError(f"Could not read {path.name}: {e}") from e


def _read_spreadsheet(file_path: str, name: str) -> list[str]:
    """Read a workbook into one tab-separated line per row.

    Sheet titles are emitted as their own chunk so a truncated workbook still
    shows which sheet the surviving rows came from.

    data_only=True asks for the last values Excel cached rather than formula
    text. A workbook written by a script and never opened in Excel has no
    cached values, so formula cells read as empty — that is a property of the
    file, not a bug here, and is worth knowing before trusting a blank column.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError(
            f"Cannot read {name}: openpyxl is not installed. "
            "Install it with: pip install openpyxl"
        ) from e

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read {name}: {e}") from e

    chunks: list[str] = []
    try:
        for sheet in workbook.worksheets:
            chunks.append(f"### Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                if row is None or all(cell is None for cell in row):
                    continue  # blank spacer rows carry nothing
                chunks.append(
                    "\t".join("" if cell is None else str(cell) for cell in row)
                )
    finally:
        workbook.close()
    return chunks


def _run_extractor(cmd: list[str], name: str) -> str:
    """Run a text-extraction command, turning its failures into ValueError."""
    try:
        return subprocess.run(
            cmd, capture_output=True, text=True, timeout=120, check=True
        ).stdout
    except subprocess.CalledProcessError as e:
        raise ValueError(f"Could not read {name}: {e.stderr.strip()[:200]}") from e
    except subprocess.TimeoutExpired as e:
        raise ValueError(f"Timed out reading {name}") from e


def extract_document_text(file_path: str) -> str:
    """Extract a document's text as a labelled text content part.

    Trims to MAX_DOCUMENT_TOKENS on a chunk boundary and says how much it kept,
    because a document silently cut in half produces a confident answer about
    the half that was sent.

    Raises:
        ValueError: If the file cannot be read or has no extractable text.
    """
    name = Path(file_path).name
    chunks, unit = _read_document(file_path)

    if not chunks:
        raise ValueError(f"{name} contains no extractable text.")

    total_chars = sum(len(c) for c in chunks)
    if unit == "page" and total_chars / len(chunks) < MIN_CHARS_PER_PAGE:
        raise ValueError(
            f"{name} yielded only {total_chars} characters across {len(chunks)} "
            "pages, which means it is images of text rather than text. That "
            "needs OCR — this UI reads text layers only."
        )

    budget = MAX_DOCUMENT_TOKENS * CHARS_PER_TOKEN
    kept: list[str] = []
    used = 0
    for chunk in chunks:
        if used + len(chunk) > budget:
            break
        kept.append(chunk)
        used += len(chunk)

    plural = "" if len(chunks) == 1 else "s"
    if not kept:  # a single chunk larger than the whole budget
        kept = [chunks[0][:budget]]
        header = (
            f"[Attached document: {name} — first {unit} of {len(chunks)}, "
            "truncated to fit the context window]"
        )
    elif len(kept) < len(chunks):
        header = (
            f"[Attached document: {name} — {unit}s 1-{len(kept)} of "
            f"{len(chunks)}; the remaining {len(chunks) - len(kept)} did not "
            "fit in the context window and were NOT sent]"
        )
    else:
        header = (
            f"[Attached document: {name} — {len(chunks)} {unit}{plural}, "
            f"~{used // CHARS_PER_TOKEN} tokens]"
        )

    print(f"[Chat] {header}", flush=True)
    separator = CHUNK_SEPARATORS.get(unit, "\n\n")
    return header + "\n\n" + separator.join(kept).strip()


def extract_document_figures(file_path: str) -> list[str]:
    """Extract a PDF's embedded figure images as base64 data URLs.

    Only embedded rasters are found, which is what most journals ship. A paper
    whose figures are vector art yields nothing here; the caller says so rather
    than implying the document had no figures.

    Downscales anything that would blow the token budget, largest figures first,
    and returns at most MAX_FIGURE_IMAGES.
    """
    if shutil.which("pdfimages") is None:
        return []

    from PIL import Image

    with tempfile.TemporaryDirectory() as tmp:
        prefix = str(Path(tmp) / "fig")
        try:
            subprocess.run(
                ["pdfimages", "-png", file_path, prefix],
                capture_output=True, timeout=180, check=True,
            )
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired):
            return []

        candidates = []
        for png in sorted(Path(tmp).glob("fig-*.png")):
            try:
                with Image.open(png) as im:
                    w, h = im.size
            except Exception:
                continue
            if min(w, h) < MIN_FIGURE_EDGE:
                continue  # logo, rule, or icon
            candidates.append((w * h, png))

        # Biggest first: the figure that carries the argument is rarely the
        # publisher's mark.
        candidates.sort(reverse=True)

        urls, budget = [], MAX_FIGURE_TOKENS * FIGURE_PIXELS_PER_TOKEN
        for pixels, png in candidates[:MAX_FIGURE_IMAGES]:
            with Image.open(png) as im:
                im = im.convert("RGB")
                if pixels > budget:
                    if budget < FIGURE_PIXELS_PER_TOKEN * 200:
                        break
                    scale = (budget / pixels) ** 0.5
                    im = im.resize((max(1, int(im.width * scale)),
                                    max(1, int(im.height * scale))))
                    pixels = im.width * im.height
                out = Path(tmp) / f"send-{png.stem}.png"
                im.save(out, "PNG")
            budget -= pixels
            data = base64.b64encode(out.read_bytes()).decode("utf-8")
            urls.append(f"data:image/png;base64,{data}")
        return urls


def encode_file_to_base64(file_path: str) -> tuple[str, str]:
    """
    Encode a file to base64 data URL.

    Resolution order: known extension, then the system MIME database, then
    an error. Guessing wrong here is worse than failing: labelling a video
    "image/jpeg" makes the server reject or misread it, and labelling
    anything unknown as an image is how an attachment ends up silently
    dropped instead of reported.

    Returns:
        Tuple of (data_url, media_type) where media_type is 'image' or 'video'

    Raises:
        ValueError: If the file type cannot be resolved to an image or video.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix in IMAGE_TYPES:
        mime_type, media_type = IMAGE_TYPES[suffix], "image"
    elif suffix in VIDEO_TYPES:
        mime_type, media_type = VIDEO_TYPES[suffix], "video"
    else:
        guessed, _ = mimetypes.guess_type(file_path)
        if guessed and guessed.startswith("image/"):
            mime_type, media_type = guessed, "image"
        elif guessed and guessed.startswith("video/"):
            mime_type, media_type = guessed, "video"
        else:
            raise ValueError(
                f"Unsupported file type {suffix or path.name!r} "
                f"(detected MIME: {guessed or 'unknown'}). "
                "Supported: "
                + ", ".join(
                    sorted(IMAGE_TYPES) + sorted(VIDEO_TYPES) + sorted(DOCUMENT_TYPES)
                )
            )

    with open(file_path, "rb") as f:
        data = base64.b64encode(f.read()).decode("utf-8")

    return f"data:{mime_type};base64,{data}", media_type


def build_media_items(files: list[str]) -> list[dict]:
    """Turn attachments into OpenAI-format content parts.

    Images and video become media parts; documents become text parts, because
    the model has no document input and a PDF's text layer is better read than
    rendered. Every file produces a part or an exception — no file is skipped,
    because a skipped attachment is invisible to the user and to the server.
    """
    items = []
    for file_path in files:
        suffix = Path(file_path).suffix.lower()
        if suffix in DOCUMENT_TYPES:
            # Only a PDF can carry embedded figures; .txt and .docx are text.
            figures = extract_document_figures(file_path) if suffix == ".pdf" else []
            if figures:
                note = f" {len(figures)} figure image(s) attached."
            elif suffix == ".pdf":
                note = (
                    " No embedded figure images found — any figures in this PDF "
                    "are vector art and were NOT sent, so do not describe them."
                )
            else:
                note = " This is a text document with no images."
            items.append(
                {"type": "text", "text": extract_document_text(file_path) + "\n" + note}
            )
            for url in figures:
                items.append({"type": "image_url", "image_url": {"url": url}})
            if suffix == ".pdf":
                print(f"[Chat] PDF figures attached: {len(figures)}", flush=True)
            continue
        data_url, media_type = encode_file_to_base64(file_path)
        key = "image_url" if media_type == "image" else "video_url"
        items.append({"type": key, key: {"url": data_url}})
    return items


def build_message_content(
    text: str,
    files: list[str] | None = None,
    media_items: list[dict] | None = None,
) -> list | str:
    """
    Build OpenAI-compatible message content with text and optional files.

    Args:
        text: The text message
        files: Optional list of file paths (images or videos)
        media_items: Optional pre-encoded media parts, to avoid base64-encoding
            the same (possibly very large) file twice in one turn

    Returns:
        Content in OpenAI multimodal format
    """
    if not files and not media_items:
        return text

    content = []

    # Add text part first
    if text:
        content.append({"type": "text", "text": text})

    if media_items is None:
        media_items = build_media_items(files or [])
    content.extend(media_items)

    return content if content else text


# A single image or a sampled video costs the model hundreds to thousands of
# prompt tokens. If media was attached and the whole prompt came in under this,
# the attachment never reached the model.
MIN_PROMPT_TOKENS_WITH_MEDIA = 200

# Historical image/video bytes are intentionally not replayed on every turn.
# Reattaching is explicit and avoids repeatedly decoding and prefilling the same
# large media payload for ordinary text follow-ups.
OMITTED_MEDIA_NOTE = (
    "[A media attachment from this earlier turn is not included in this request. "
    "Ask the user to reattach it before making new claims about its visual content.]"
)


def media_drop_warning(media_items: list[dict], prompt_tokens: int) -> str | None:
    """Detect an attachment that was accepted and then silently dropped.

    Applies to image and video attachments only; see the call site for why a
    text document is exempt.

    Every media-handling bug found in this stack returned HTTP 200 with a
    fluent answer — "I don't see any video" — and everyone blamed the model.
    The cheapest reliable discriminator needs no ground truth: compare the
    prompt token count against the same request with no attachment. A dropped
    attachment leaves it in the dozens; a real one puts it in the thousands.
    """
    if not media_items or not prompt_tokens:
        return None
    if prompt_tokens >= MIN_PROMPT_TOKENS_WITH_MEDIA:
        return None
    kinds = ", ".join(sorted({item["type"].replace("_url", "") for item in media_items}))
    return (
        f"WARNING: {len(media_items)} attachment(s) ({kinds}) were sent, but the "
        f"server counted only {prompt_tokens} prompt tokens - far too few to "
        "contain them. The attachment was almost certainly dropped before the "
        "model saw it, so the answer above describes nothing. Check that the "
        "server loaded a multimodal model (GET /health -> model_type should be "
        '"mllm"; force it with `vllm-mlx serve --mllm ...`).'
    )


def check_server_multimodal(server_url: str) -> str | None:
    """Return a warning if the server is not serving a multimodal model.

    Model capability is detected from the model's config.json, but a text-only
    load accepts image/video requests anyway and silently discards the media
    parts. Better to say so at startup than to let every answer look like a
    model that cannot see.
    """
    try:
        response = requests.get(f"{server_url}/health", timeout=10)
        response.raise_for_status()
        health = response.json()
    except Exception as e:
        return f"Could not reach {server_url}/health ({e}); is the server running?"

    if not health.get("model_loaded", False):
        return None  # lazy load: capability is not known yet

    if health.get("model_type") != "mllm":
        return (
            f"Server model {health.get('model_name')!r} is loaded as text-only "
            f"(model_type={health.get('model_type')!r}). Images and videos will "
            "be dropped from every request without an error. Restart the server "
            "with a multimodal model, or force it with `vllm-mlx serve --mllm`."
        )
    return None


def split_reasoning(message: dict) -> tuple[str, str]:
    """Separate a reply into (reasoning, answer).

    Prefers the server's own split: with --reasoning-parser, the thinking
    arrives in reasoning_content and content holds only the answer. Falls back
    to splitting a raw <think> block, so the UI behaves the same whether or not
    the parser is enabled — and an unterminated block (the model hit the token
    ceiling mid-thought) is shown as reasoning rather than swallowed.
    """
    answer = message.get("content") or ""
    reasoning = message.get("reasoning_content") or ""

    if not reasoning and "<think>" in answer:
        before, _, rest = answer.partition("<think>")
        thought, closed, after = rest.partition("</think>")
        reasoning = thought.strip()
        answer = (before + after).strip() if closed else before.strip()

    return reasoning.strip(), answer.strip()


def create_chat_function(
    server_url: str,
    max_tokens: int,
    temperature: float,
    served_model_name: str = "default",
):
    """
    Create the chat function for Gradio ChatInterface.

    Args:
        server_url: URL of the vllm-mlx server
        max_tokens: Maximum tokens to generate
        temperature: Sampling temperature
        served_model_name: Model name to send in OpenAI-compatible requests

    Returns:
        Chat function compatible with gr.ChatInterface
    """
    # Remember which historical user messages had media without retaining or
    # replaying their base64 payloads. A 16 MB video becomes roughly 22 MB in a
    # request and otherwise gets decoded and prefetched again on every turn.
    media_message_indexes: set[int] = set()

    def chat(message: dict, history: list) -> str:
        """
        Process a multimodal message and return response.

        Args:
            message: Dict with 'text' and optional 'files' keys
            history: List of previous messages

        Returns:
            Assistant response text
        """
        # Extract text and files from message
        text = message.get("text", "") if isinstance(message, dict) else message
        files = message.get("files", []) if isinstance(message, dict) else []

        # Gradio reuses the chat function after the user clears the conversation.
        # Drop indexes from the previous conversation so they cannot annotate a
        # new chat whose message numbering starts at zero again.
        if not history:
            media_message_indexes.clear()

        # Debug output
        import sys

        print(f"[Chat] Text: {text!r}", flush=True)
        print(f"[Chat] Files: {files}", flush=True)
        print(f"[Chat] History length: {len(history)}", flush=True)
        sys.stdout.flush()

        # Build messages list for API
        messages = []

        omitted_media_count = 0
        # Process history as text only. Historical media is represented by a
        # marker so the model knows it must request a reattachment rather than
        # hallucinating unseen visual details.
        for i, msg in enumerate(history):
            if isinstance(msg, dict):
                role = msg.get("role", "user")
                content = msg.get("content", "")

                if isinstance(content, list):
                    text_parts = [
                        p.get("text", "")
                        for p in content
                        if isinstance(p, dict) and p.get("type") == "text"
                    ]
                    content = " ".join(text_parts)
                elif isinstance(content, dict):
                    content = content.get("text", str(content))
                elif not isinstance(content, str):
                    content = str(content)

                if i in media_message_indexes and role == "user":
                    content = (
                        f"{content}\n\n{OMITTED_MEDIA_NOTE}"
                        if content
                        else OMITTED_MEDIA_NOTE
                    )
                    omitted_media_count += 1

                messages.append({"role": role, "content": content})

        if omitted_media_count:
            print(
                f"[Chat] Omitted historical media from {omitted_media_count} "
                "message(s); reattach to analyze it again",
                flush=True,
            )

        # Encode attachments once: a phone video is tens of megabytes and
        # base64-encoding it twice per turn is pure latency.
        try:
            media_items = build_media_items(files) if files else []
        except ValueError as e:
            return f"Error: {e}"

        # Build current message content. Media is sent only on this upload turn.
        current_content = build_message_content(
            text, files if files else None, media_items=media_items
        )
        messages.append({"role": "user", "content": current_content})

        # Remember attachment presence, not its base64 bytes. This message will
        # appear at index len(history) when Gradio submits the next turn.
        if media_items:
            current_idx = len(history)
            media_message_indexes.add(current_idx)
            print(
                f"[Chat] Sent {len(media_items)} media item(s) for message {current_idx}; "
                "historical replay disabled",
                flush=True,
            )

        # Debug
        print(f"[Chat] Sending {len(messages)} messages to server")
        if isinstance(current_content, list):
            print(f"[Chat] Content types: {[c.get('type') for c in current_content]}")

        # Send request to server
        try:
            response = requests.post(
                f"{server_url}/v1/chat/completions",
                json={
                    "model": served_model_name,
                    "messages": messages,
                    "max_tokens": max_tokens,
                    "temperature": temperature,
                },
                timeout=600,
            )
            response.raise_for_status()
            result = response.json()
            reasoning, answer = split_reasoning(
                result["choices"][0]["message"]
            )

            usage = result.get("usage", {}) or {}
            prompt_tokens = usage.get("prompt_tokens", 0)
            print(
                f"[Chat] prompt_tokens={prompt_tokens} "
                f"reasoning_chars={len(reasoning)}",
                flush=True,
            )

            # Only image and video parts can vanish without trace. A
            # document arrives as text, so it is in the prompt by
            # construction and its token cost is proportional to its
            # size — a small spreadsheet legitimately costs ~50 tokens,
            # which would trip the threshold and cry wolf.
            sent_media = [
                item for item in media_items if item["type"] != "text"
            ]
            warning = media_drop_warning(sent_media, prompt_tokens)
            if warning:
                print(f"[Chat] {warning}", flush=True)
                answer = f"{answer}\n\n---\n{warning}"

            if not answer and reasoning:
                # The token budget ran out inside the thinking block. Say so,
                # rather than returning an empty bubble.
                answer = (
                    "_The reply hit the token limit while still reasoning, so "
                    "there is no final answer. Open the reasoning below, or "
                    "raise --max-tokens._"
                )

            if not reasoning:
                return answer

            # metadata={"title": ...} is what makes Gradio render a message as a
            # collapsed thought block instead of prose in the transcript.
            return [
                gr.ChatMessage(
                    role="assistant",
                    content=reasoning,
                    metadata={"title": "Reasoning", "status": "done"},
                ),
                gr.ChatMessage(role="assistant", content=answer),
            ]

        except requests.exceptions.ConnectionError:
            return "Error: Cannot connect to server. Make sure vllm-mlx is running."
        except requests.exceptions.Timeout:
            return "Error: Timeout - server took too long to respond."
        except Exception as e:
            return f"Error: {str(e)}"

    return chat


def main():
    """Run the Gradio app."""
    parser = argparse.ArgumentParser(
        description="Gradio Multimodal Chat Interface for vllm-mlx",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Start with default settings
    vllm-mlx-chat

    # Connect to a different server
    vllm-mlx-chat --server-url http://localhost:9000

    # Create a public share link
    vllm-mlx-chat --share

Note: Make sure the vllm-mlx server is running with a multimodal model:
    vllm-mlx serve --served-model-name default mlx-community/Qwen3-VL-4B-Instruct-3bit --port 8000
        """,
    )
    parser.add_argument(
        "--server-url",
        type=str,
        default="http://localhost:8000",
        help="URL of the vllm-mlx server (default: http://localhost:8000)",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=7860,
        help="Port for Gradio interface (default: 7860)",
    )
    parser.add_argument(
        "--share",
        action="store_true",
        help="Create a public share link",
    )
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=2048,
        help="Maximum tokens to generate (default: 2048)",
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=0.7,
        help="Sampling temperature (default: 0.7)",
    )
    parser.add_argument(
        "--served-model-name",
        type=str,
        default="default",
        help=(
            "Model name to send in /v1/chat/completions requests " "(default: default)"
        ),
    )
    parser.add_argument(
        "--text-only",
        action="store_true",
        help="Use text-only mode (no image/video support, faster for LLM-only models)",
    )
    args = parser.parse_args()

    print(f"Connecting to vllm-mlx server at: {args.server_url}")
    print(f"Starting Gradio interface on port: {args.port}")

    if args.text_only:
        print("Mode: Text-only (no multimodal support)")

        # Create text-only chat function
        def text_chat(message: str, history: list) -> str:
            """Process a text-only message."""
            messages = []
            for msg in history:
                if isinstance(msg, dict):
                    role = msg.get("role", "user")
                    content = msg.get("content", "")
                    if isinstance(content, list):
                        text_parts = [
                            p.get("text", "")
                            for p in content
                            if isinstance(p, dict) and p.get("type") == "text"
                        ]
                        content = " ".join(text_parts)
                    messages.append({"role": role, "content": content})

            messages.append({"role": "user", "content": message})

            try:
                response = requests.post(
                    f"{args.server_url}/v1/chat/completions",
                    json={
                        "model": args.served_model_name,
                        "messages": messages,
                        "max_tokens": args.max_tokens,
                        "temperature": args.temperature,
                    },
                    timeout=120,
                )
                response.raise_for_status()
                result = response.json()
                return result["choices"][0]["message"]["content"]
            except requests.exceptions.ConnectionError:
                return "Error: Cannot connect to server. Make sure vllm-mlx is running."
            except requests.exceptions.Timeout:
                return "Error: Timeout - server took too long to respond."
            except Exception as e:
                return f"Error: {str(e)}"

        demo = gr.ChatInterface(
            fn=text_chat,
            title="vllm-mlx Text Chat",
            description="Fast text-only chat with LLM models on Apple Silicon.",
            examples=[
                "Hello, who are you?",
                "Explain quantum computing in simple terms.",
                "Write a haiku about programming.",
            ],
        )
    else:
        print("Mode: Multimodal (text, image, video, and documents)")

        capability_warning = check_server_multimodal(args.server_url)
        if capability_warning:
            print(f"WARNING: {capability_warning}")

        # Create chat function
        chat_fn = create_chat_function(
            server_url=args.server_url,
            max_tokens=args.max_tokens,
            temperature=args.temperature,
            served_model_name=args.served_model_name,
        )

        # Create ChatInterface with multimodal support
        description = (
            "Chat with vision-language models on Apple Silicon. "
            "Upload images or videos! Attachments are analyzed only on the turn "
            "where they are uploaded; reattach one for later visual questions."
        )
        if capability_warning:
            description += f"\n\n**WARNING:** {capability_warning}"

        demo = gr.ChatInterface(
            fn=chat_fn,
            title="vllm-mlx Multimodal Chat",
            description=description,
            multimodal=True,
            textbox=gr.MultimodalTextbox(
                file_types=UPLOAD_FILE_TYPES,
                file_count="multiple",
                placeholder="Type a message or upload an image/video...",
            ),
        )

    demo.launch(
        server_port=args.port,
        share=args.share,
    )


if __name__ == "__main__":
    main()
